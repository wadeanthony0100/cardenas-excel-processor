"""
title: heatChart.py
author: Wade Mauger
date: 4/22/2015
description:Utilizes the openpyxl library
to open and process an Excel file and
output an appropriate Excel visual file.

"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import fills, Color, Fill, Style, PatternFill
from openpyxl.styles.colors import BLACK

#define how a filled cell is to be filled
from openpyxl.utils import get_column_letter

aFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='6F0701')
bFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='872712')
cFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='9D4724')
dFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='B76736')
eFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='CF8747')
fFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='E7A759')
gFill = PatternFill(patternType=fills.FILL_SOLID, fgColor='FFC76B')
backFill = PatternFill(patternType=fills.FILL_SOLID, fgColor=BLACK)

"""
Cells will be colored in based on frequence of their correspondance
as follows:

1-2:    A
2-4:    B
4-8:    C
8-16:   D
16-32:  E
32-64:  F
64+: G

"""


class Correlation():
    def __init__(self, meaning, arrow_type):
        self.meaning = meaning
        self.arrow_type = int(arrow_type)
        self.frequency = 1

    def __repr__(self):
        return "Corr(meaning="+str(self.meaning)+", arrow_type="+str(self.arrow_type)+", frequency="+str(self.frequency)+" )"

    def __eq__(self, other):
        return self.arrow_type == other.arrow_type

    def addFreq(self):
        """
        add one to this instance of a correlation
        :return: Nonetype
        """
        self.frequency += 1


#collect I/O data
input_name = input("Input file name- This must be placed \nin the directory above this script: ")
output_name = input("Output file name: ")
vertical_letter = input("Which column of the input file is meant to be the horizontal axis? ").upper()
horizontal_letter = input("Which column of the input file is meant to be the vertical axis? ").upper()

#ensure proper formatting
if input_name[-5:] != '.xlsx':
    input_name += '.xlsx'

if output_name[-5:] != '.xlsx':
    output_name += '.xlsx'

#initialize workbook, get data worksheet
wb = load_workbook(filename = input_name,  use_iterators=True)
data_sheet = wb.get_sheet_by_name('Data')

#get the number of valid rows in the sheet, for iteration
num_rows = data_sheet.max_row

#initialize a dictionary of meaning strings
#data will be stored as a mapping of meaning string keys
#to lists of arrows for which the meaning exists
meanings_to_lst_arrows = {}
meanings_to_lst_corr_objs = {}
max_arrow = 0

for x in range(2, num_rows+1):
    arrow = data_sheet[vertical_letter + str(x)].value
    meaning = data_sheet[horizontal_letter + str(x)].value
    if meaning is not None:
        meaning = meaning.strip()
    if arrow > max_arrow:
        max_arrow = arrow
    if meaning in meanings_to_lst_arrows:
        if not arrow in meanings_to_lst_arrows[meaning]:
            meanings_to_lst_arrows[meaning].append(arrow)
        isIn = False
        for corr in meanings_to_lst_corr_objs[meaning]:
            if corr == Correlation("", arrow):
                isIn = True
                corr.addFreq()
        if isIn == False:
            new_corr = Correlation(meaning, arrow)
            meanings_to_lst_corr_objs[meaning].append(new_corr)
    else:
        meanings_to_lst_arrows[meaning] = [arrow]
        new_corr = Correlation(meaning, arrow)
        meanings_to_lst_corr_objs[meaning] = [new_corr]

print(meanings_to_lst_corr_objs)

#make a new Worksheet
new_book = Workbook()
new_sheet = new_book.active

#label Axies and Data
for x in range(1, max_arrow+1):
    new_sheet[get_column_letter(x + 1) + "1"] = x

#style all spaces black
for x in range(2, len(meanings_to_lst_arrows) + 1):
    for y in range(2, max_arrow + 2):
        new_sheet[str(get_column_letter(y)) + str(x)].style = Style(fill=backFill)

#get an alphabetized list to iterate through
keys_list = list(meanings_to_lst_arrows.keys())
keys_list = [x for x in keys_list if x != None]
keys_list.sort()
index = 2
for key in keys_list:
    new_sheet["A" + str(index)] = key
    for corr in meanings_to_lst_corr_objs[key]:
        tier_num = corr.frequency
        print("Meaning = " +str(corr.meaning)+ ",\tTier num = " + str(tier_num))
        if tier_num <= 2:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=aFill)

        elif tier_num <= 4:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=bFill)

        elif tier_num <= 8:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=cFill)

        elif tier_num <= 16:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=dFill)

        elif tier_num <= 32:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=eFill)

        elif tier_num <= 64:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=fFill)

        else:
            new_sheet[str(get_column_letter(corr.arrow_type + 1)) + str(index)].style = Style(fill=gFill)

    index += 1
"""
#test the fill colors
new_sheet["A1"].style = Style(fill=aFill)
new_sheet["B1"].style = Style(fill=bFill)
new_sheet["C1"].style = Style(fill=cFill)
new_sheet["D1"].style = Style(fill=dFill)
new_sheet["E1"].style = Style(fill=eFill)
new_sheet["F1"].style = Style(fill=fFill)
new_sheet["G1"].style = Style(fill=gFill)
new_sheet["H1"].style = Style(fill=hFill)
"""
#save the new workbook with the given output filename
new_book.save(output_name)